"""Excel format converter for DocFlow."""

from typing import List, Optional, Union

from ..types import ConvertResult


class ExcelConverter:
    """Converts between Excel (XLSX) and Markdown formats.
    
    Requires openpyxl: pip install openpyxl
    
    Example:
        >>> converter = ExcelConverter()
        >>> md = converter.to_markdown(excel_bytes, "report.xlsx")
    """

    def __init__(
        self,
        sheet_name: Optional[str] = None,
        include_all_sheets: bool = True,
    ) -> None:
        """Initialize Excel converter.
        
        Args:
            sheet_name: Specific sheet to convert (None = first or all).
            include_all_sheets: Include all sheets in markdown.
        """
        self.sheet_name = sheet_name
        self.include_all_sheets = include_all_sheets

    def to_markdown(
        self,
        excel_data: bytes,
        filename: str = "data.xlsx",
        include_metadata: bool = True,
    ) -> ConvertResult:
        """Convert Excel to Markdown.
        
        Args:
            excel_data: Excel file bytes.
            filename: Original filename.
            include_metadata: Include YAML frontmatter.
            
        Returns:
            ConvertResult with markdown content.
        """
        try:
            from openpyxl import load_workbook
            from io import BytesIO
        except ImportError:
            return ConvertResult(
                success=False,
                error="openpyxl is required: pip install openpyxl"
            )

        try:
            wb = load_workbook(BytesIO(excel_data), data_only=True)
            md_parts = []

            # Metadata
            if include_metadata:
                md_parts.append("---")
                md_parts.append(f"source: {filename}")
                md_parts.append("format: xlsx")
                md_parts.append(f"sheets: {len(wb.sheetnames)}")
                md_parts.append(f"sheet_names: {wb.sheetnames}")
                md_parts.append("---\n")

            # Title
            title = filename.replace(".xlsx", "").replace(".xls", "").replace("_", " ").title()
            md_parts.append(f"# {title}\n")

            # Process sheets
            sheets_to_process = []
            if self.sheet_name:
                if self.sheet_name in wb.sheetnames:
                    sheets_to_process = [self.sheet_name]
                else:
                    return ConvertResult(
                        success=False,
                        error=f"Sheet '{self.sheet_name}' not found"
                    )
            elif self.include_all_sheets:
                sheets_to_process = wb.sheetnames
            else:
                sheets_to_process = [wb.sheetnames[0]]

            total_rows = 0
            for sheet_name in sheets_to_process:
                ws = wb[sheet_name]
                
                if len(sheets_to_process) > 1:
                    md_parts.append(f"\n## {sheet_name}\n")

                # Get data
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    md_parts.append("*Empty sheet*\n")
                    continue

                # Filter out completely empty rows
                rows = [r for r in rows if any(c is not None for c in r)]
                if not rows:
                    md_parts.append("*Empty sheet*\n")
                    continue

                total_rows += len(rows)
                max_cols = max(len(r) for r in rows)

                # Header
                header = rows[0]
                header = [str(c) if c is not None else "" for c in header]
                header = header + [""] * (max_cols - len(header))
                
                md_parts.append("| " + " | ".join(self._escape_cell(c) for c in header) + " |")
                md_parts.append("| " + " | ".join("---" for _ in header) + " |")

                # Data rows
                for row in rows[1:]:
                    cells = [str(c) if c is not None else "" for c in row]
                    cells = cells + [""] * (max_cols - len(cells))
                    md_parts.append("| " + " | ".join(self._escape_cell(c) for c in cells) + " |")

                md_parts.append("")

            return ConvertResult(
                success=True,
                content="\n".join(md_parts),
                format="xlsx",
                metadata={
                    "sheets": len(sheets_to_process),
                    "total_rows": total_rows,
                },
            )

        except Exception as e:
            return ConvertResult(success=False, error=str(e))

    def from_markdown(
        self,
        markdown: str,
        sheet_name: str = "Sheet1",
    ) -> ConvertResult:
        """Convert Markdown tables to Excel.
        
        Args:
            markdown: Markdown content with tables.
            sheet_name: Name for the Excel sheet.
            
        Returns:
            ConvertResult with Excel bytes.
        """
        try:
            from openpyxl import Workbook
            from io import BytesIO
        except ImportError:
            return ConvertResult(
                success=False,
                error="openpyxl is required: pip install openpyxl"
            )

        try:
            tables = self._extract_tables(markdown)
            
            if not tables:
                return ConvertResult(success=False, error="No tables found in markdown")

            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Write first table
            for row_idx, row in enumerate(tables[0], 1):
                for col_idx, cell in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell)

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return ConvertResult(
                success=True,
                content=output.getvalue(),
                format="xlsx",
                metadata={"tables_converted": 1},
            )

        except Exception as e:
            return ConvertResult(success=False, error=str(e))

    def _extract_tables(self, markdown: str) -> List[List[List[str]]]:
        """Extract tables from markdown."""
        tables = []
        lines = markdown.split("\n")
        
        current_table = []
        in_table = False
        
        for line in lines:
            line = line.strip()
            
            if line.startswith("|") and line.endswith("|"):
                if set(line.replace("|", "").replace("-", "").replace(":", "").strip()) == set():
                    continue
                
                cells = [c.strip() for c in line.split("|")[1:-1]]
                current_table.append(cells)
                in_table = True
            else:
                if in_table and current_table:
                    tables.append(current_table)
                    current_table = []
                in_table = False
        
        if current_table:
            tables.append(current_table)
        
        return tables

    def _escape_cell(self, cell: str) -> str:
        """Escape cell content for markdown."""
        return str(cell).replace("|", "\\|").replace("\n", " ")
